from io import BytesIO

from openpyxl import Workbook

from services.campaign_spreadsheet import parse_campaign_workbook


def _workbook_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Camil"
    sheet.cell(6, 3, "RCA")
    sheet.cell(6, 7, "Faturado")
    sheet.cell(6, 8, "Objetivo")
    sheet.cell(6, 9, "Realizado")
    sheet.cell(7, 3, 607)
    sheet.cell(7, 6, 1000)
    sheet.cell(7, 7, 1234.5)
    sheet.cell(7, 8, 20)
    sheet.cell(7, 9, 22)

    awarded = workbook.create_sheet("Itens Premiados")
    awarded.cell(4, 10, "2026-07-14")
    awarded.cell(6, 5, "Baston")
    awarded.cell(7, 6, 120000)
    awarded.cell(8, 6, 600000)
    awarded.cell(9, 1, "RCA")
    awarded.cell(10, 1, 607)
    awarded.cell(10, 2, "JOSE")
    awarded.cell(10, 4, 50)
    awarded.cell(10, 5, 50)
    awarded.cell(10, 6, 900)
    awarded.cell(10, 7, 10)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_parser_extrai_meta_geral_e_resultado_individual():
    result = parse_campaign_workbook(_workbook_bytes())

    assert result["suppliers"] == [{
        "name": "Baston",
        "goal": 600000.0,
        "realized": 120000.0,
        "remaining": 480000.0,
        "percentage": 20.0,
    }]
    rca = result["rcaResults"]["607"]
    assert rca["totalPrize"] == 50
    assert rca["awardedItems"][0]["boxes"] == 10
    assert rca["industries"]["Camil"] == {
        "minimumSales": 1000.0,
        "sales": 1234.5,
        "targetQuantity": 20.0,
        "quantity": 22.0,
    }


def test_parser_ignora_bloco_na_quebrado():
    content = _workbook_bytes()
    result = parse_campaign_workbook(content)

    assert [item["name"] for item in result["suppliers"]] == ["Baston"]
