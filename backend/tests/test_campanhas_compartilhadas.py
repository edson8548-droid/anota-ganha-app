import os
import sys
import asyncio
from datetime import datetime
from io import BytesIO

from bson import ObjectId
from fastapi import FastAPI
from fastapi import HTTPException
from fastapi.security import HTTPAuthorizationCredentials
from fastapi.testclient import TestClient
from openpyxl import Workbook

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from routes import campanhas_compartilhadas as cc


class _Cursor:
    def __init__(self, docs):
        self._docs = list(docs)

    def sort(self, *a, **k):
        return self

    def __aiter__(self):
        self._it = iter(self._docs)
        return self

    async def __anext__(self):
        try:
            return next(self._it)
        except StopIteration:
            raise StopAsyncIteration


class _Collection:
    def __init__(self):
        self.docs = []

    def _match(self, doc, flt):
        for k, v in flt.items():
            if isinstance(v, dict) and "$in" in v:
                if doc.get(k) not in v["$in"]:
                    return False
            elif doc.get(k) != v:
                return False
        return True

    def find(self, flt=None):
        flt = flt or {}
        return _Cursor([d for d in self.docs if self._match(d, flt)])

    async def find_one(self, flt):
        for d in self.docs:
            if self._match(d, flt):
                return d
        return None

    async def insert_one(self, doc):
        doc.setdefault("_id", ObjectId())
        self.docs.append(doc)
        return type("R", (), {"inserted_id": doc["_id"]})()

    async def update_one(self, flt, update, upsert=False):
        for d in self.docs:
            if self._match(d, flt):
                for key, value in update.get("$set", {}).items():
                    target = d
                    parts = key.split(".")
                    for part in parts[:-1]:
                        target = target.setdefault(part, {})
                    target[parts[-1]] = value
                return type("R", (), {"matched_count": 1})()
        if upsert:
            novo = {}
            novo.update(update.get("$setOnInsert", {}))
            novo.update(update.get("$set", {}))
            await self.insert_one(novo)
        return type("R", (), {"matched_count": 0})()

    async def delete_one(self, flt):
        for i, d in enumerate(self.docs):
            if self._match(d, flt):
                del self.docs[i]
                return type("R", (), {"deleted_count": 1})()
        return type("R", (), {"deleted_count": 0})()

    async def delete_many(self, flt):
        antes = len(self.docs)
        self.docs = [d for d in self.docs if not self._match(d, flt)]
        return type("R", (), {"deleted_count": antes - len(self.docs)})()


class _FakeDB:
    def __init__(self):
        self.master_campaigns = _Collection()
        self.campaign_access = _Collection()


def _make_client(admin_uid="admin-1", rca_uid="rca-1"):
    cc.init_campanhas_compartilhadas(_FakeDB())
    app = FastAPI()
    app.include_router(cc.router, prefix="/api/cc")
    app.dependency_overrides[cc._require_admin] = lambda: admin_uid
    app.dependency_overrides[cc._rca_uid] = lambda: rca_uid
    return TestClient(app)


_MESTRE = {
    "nome": "Spani Julho 2026",
    "code": "spani2026",
    "distribuidora": "spani",
    "regulamento": "Compre e ganhe.",
    "startDate": "2026-07-01",
    "endDate": "2026-07-31",
    "industries": {
        "Mondelez": {"produtos": {
            "Bis": {"codigo": "123", "nome": "Bis", "ean": "7891234000019"},
        }},
        "Nestlé": {"produtos": {
            "Nescau": {"codigo": "456", "nome": "Nescau 2kg", "ean": "7891000432679"},
        }},
    },
}


def _criar(client, over=None):
    return client.post("/api/cc/admin/mestre", json={**_MESTRE, **(over or {})})


def _desbloquear(client, code="spani2026"):
    return client.post("/api/cc/desbloquear", json={"code": code})


def _spreadsheet_bytes(code=607, name="JOSE", period=None):
    period = period or datetime(2026, 7, 14)
    workbook = Workbook()
    workbook.active.title = "Camil"
    awarded = workbook.create_sheet("Itens Premiados")
    awarded.cell(4, 10, period)
    awarded.cell(6, 5, "Baston")
    awarded.cell(7, 6, 120000)
    awarded.cell(8, 6, 600000)
    awarded.cell(9, 1, "RCA")
    awarded.cell(10, 1, code)
    awarded.cell(10, 2, name)
    awarded.cell(10, 4, 50)
    awarded.cell(10, 5, 50)
    awarded.cell(10, 6, 900)
    awarded.cell(10, 7, 10)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_admin_cria_mestre_sem_meta_e_com_produtos():
    client = _make_client()
    r = _criar(client)
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["slug"] == "spani-julho-2026"
    assert "Mondelez" in data["industries"]
    prod = data["industries"]["Mondelez"]["produtos"]["Bis"]
    assert prod["ean"] == "7891234000019"
    # A mestre não guarda meta nenhuma
    assert "targetValue" not in data["industries"]["Mondelez"]
    assert "metaSugerida" not in prod
    # Nem o hash da senha volta pro cliente
    assert "code_hash" not in data


def test_rca_desbloqueia_e_ve_em_minhas():
    client = _make_client()
    _criar(client)
    r = _desbloquear(client)
    assert r.status_code == 200, r.text
    assert r.json()["acesso"] is True
    assert r.json()["campanha"]["rcaCode"] == ""

    minhas = client.get("/api/cc/minhas")
    assert minhas.status_code == 200
    assert len(minhas.json()) == 1
    assert minhas.json()[0]["nome"] == "Spani Julho 2026"


def test_desbloquear_senha_errada_falha():
    client = _make_client()
    _criar(client)
    assert _desbloquear(client, code="errada").status_code == 404


def test_acesso_e_permanente_sem_duplicar():
    client = _make_client()
    _criar(client)
    _desbloquear(client)
    _desbloquear(client)  # 2a vez
    assert len(cc._db.campaign_access.docs) == 1  # não duplicou
    assert len(client.get("/api/cc/minhas").json()) == 1


def test_admin_edita_mestre_reflete_para_rca():
    client = _make_client()
    mid = _criar(client).json()["id"]
    _desbloquear(client)

    # Admin adiciona indústria — RCA deve ver na hora (lê a mestre viva)
    r = client.put(f"/api/cc/admin/mestre/{mid}", json={
        "industries": {**_MESTRE["industries"], "Bauducco": {"produtos": {
            "Panetone": {"codigo": "789", "nome": "Panetone", "ean": "7890000000007"}
        }}}
    })
    assert r.status_code == 200, r.text
    minhas = client.get("/api/cc/minhas").json()
    assert "Bauducco" in minhas[0]["industries"]


def test_admin_troca_senha_invalida_a_antiga():
    client = _make_client()
    mid = _criar(client).json()["id"]
    client.put(f"/api/cc/admin/mestre/{mid}", json={"code": "novasenha"})
    assert _desbloquear(client, code="spani2026").status_code == 404
    assert _desbloquear(client, code="novasenha").status_code == 200


def test_excluir_mestre_remove_acessos():
    client = _make_client()
    mid = _criar(client).json()["id"]
    _desbloquear(client)
    assert client.delete(f"/api/cc/admin/mestre/{mid}").status_code == 200
    assert cc._db.campaign_access.docs == []
    assert client.get("/api/cc/minhas").json() == []


def test_criar_exige_senha_minima():
    client = _make_client()
    assert _criar(client, {"code": "ab"}).status_code == 400
    # Mínimo agora é 6 (proteção contra chute de senha)
    assert _criar(client, {"code": "abcde"}).status_code == 400
    assert _criar(client, {"code": "abcdef"}).status_code == 200


def test_listagem_tolera_campos_legados_e_bson():
    client = _make_client()
    mid = ObjectId()
    nested_id = ObjectId()
    cc._db.master_campaigns.docs.append({
        "_id": mid,
        "nome": "Campanha SP — Turbinado",
        "updated_at": "2026-07-18T12:00:00Z",
        "industries": {
            "Camil — Turbinado": {
                "produtos": {
                    "Arroz": {"nome": "Arroz", "referencia": nested_id},
                },
            },
        },
    })

    response = client.get("/api/cc/admin/mestre")

    assert response.status_code == 200, response.text
    data = response.json()[0]
    assert data["nome"] == "Campanha SP — Turbinado"
    assert data["updated_at"] == "2026-07-18T12:00:00Z"
    assert data["industries"]["Camil — Turbinado"]["produtos"]["Arroz"]["referencia"] == str(nested_id)


def test_admin_da_campanha_usa_token_verificado_e_allowlist(monkeypatch):
    monkeypatch.setenv("ADMIN_ALLOWED_EMAILS", "edson854_8@hotmail.com")
    monkeypatch.setattr(cc.firebase_auth, "verify_id_token", lambda _token: {
        "uid": "admin-1",
        "email": "edson854_8@hotmail.com",
        "email_verified": True,
    })

    uid = asyncio.run(cc._require_admin(HTTPAuthorizationCredentials(
        scheme="Bearer",
        credentials="token-valido",
    )))

    assert uid == "admin-1"


def test_admin_da_campanha_aceita_token_legado_sem_email_verified(monkeypatch):
    monkeypatch.setenv("ADMIN_ALLOWED_EMAILS", "edson854_8@hotmail.com")
    monkeypatch.setattr(cc.firebase_auth, "verify_id_token", lambda _token: {
        "uid": "admin-1",
        "email": "edson854_8@hotmail.com",
    })

    uid = asyncio.run(cc._require_admin(HTTPAuthorizationCredentials(
        scheme="Bearer",
        credentials="token-legado-valido",
    )))

    assert uid == "admin-1"


def test_admin_da_campanha_bloqueia_email_fora_da_allowlist(monkeypatch):
    monkeypatch.setenv("ADMIN_ALLOWED_EMAILS", "edson854_8@hotmail.com")
    monkeypatch.setattr(cc.firebase_auth, "verify_id_token", lambda _token: {
        "uid": "outro-1",
        "email": "outro@example.com",
        "email_verified": True,
    })

    try:
        asyncio.run(cc._require_admin(HTTPAuthorizationCredentials(
            scheme="Bearer",
            credentials="token-valido",
        )))
    except HTTPException as exc:
        assert exc.status_code == 403
    else:
        raise AssertionError("E-mail fora da allowlist deveria ser bloqueado")


def test_listagem_admin_mostra_quantos_rcas_desbloquearam():
    client = _make_client()
    _criar(client)
    _criar(client, {"nome": "Outra Campanha", "code": "outra2026"})

    lista_antes = client.get("/api/cc/admin/mestre").json()
    assert all(item["desbloqueios"] == 0 for item in lista_antes)

    _desbloquear(client)

    lista = client.get("/api/cc/admin/mestre").json()
    por_nome = {item["nome"]: item["desbloqueios"] for item in lista}
    assert por_nome["Spani Julho 2026"] == 1
    assert por_nome["Outra Campanha"] == 0


def test_codigo_rca_nao_e_exigido_ao_desbloquear():
    client = _make_client()
    _criar(client)

    assert _desbloquear(client).status_code == 200


def test_primeiro_upload_mostra_previa_e_depois_vincula_codigo():
    client = _make_client()
    mid = _criar(client).json()["id"]
    _desbloquear(client)
    file = {"arquivo": ("apuracao.xlsx", _spreadsheet_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}

    preview = client.post(
        f"/api/cc/mestre/{mid}/apuracao",
        files=file,
        data={"rca_code": "607", "confirmar": "false"},
    )
    assert preview.status_code == 200, preview.text
    assert preview.json()["requiresConfirmation"] is True
    assert preview.json()["profile"] == {"code": "607", "name": "JOSE"}
    assert cc._db.campaign_access.docs[0].get("rca_code") is None

    confirmed = client.post(
        f"/api/cc/mestre/{mid}/apuracao",
        files=file,
        data={"rca_code": "607", "confirmar": "true"},
    )
    assert confirmed.status_code == 200, confirmed.text
    assert cc._db.campaign_access.docs[0]["rca_code"] == "607"

    minhas = client.get("/api/cc/minhas").json()
    assert minhas[0]["rcaCode"] == "607"
    assert minhas[0]["rcaResult"]["name"] == "JOSE"


def test_primeiro_upload_rejeita_codigo_ausente_na_planilha():
    client = _make_client()
    mid = _criar(client).json()["id"]
    _desbloquear(client)

    response = client.post(
        f"/api/cc/mestre/{mid}/apuracao",
        files={"arquivo": ("apuracao.xlsx", _spreadsheet_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"rca_code": "614", "confirmar": "false"},
    )

    assert response.status_code == 404
    assert "614" in response.json()["detail"]


def test_upload_rejeita_codigo_ja_vinculado_a_outra_conta():
    client = _make_client(rca_uid="rca-1")
    mid = _criar(client).json()["id"]
    _desbloquear(client)
    file = {"arquivo": ("apuracao.xlsx", _spreadsheet_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    assert client.post(
        f"/api/cc/mestre/{mid}/apuracao", files=file,
        data={"rca_code": "607", "confirmar": "true"},
    ).status_code == 200

    client.app.dependency_overrides[cc._rca_uid] = lambda: "rca-2"
    _desbloquear(client)
    conflict = client.post(
        f"/api/cc/mestre/{mid}/apuracao", files=file,
        data={"rca_code": "607", "confirmar": "true"},
    )

    assert conflict.status_code == 409
    assert "outra conta" in conflict.json()["detail"]


def test_upload_nao_substitui_apuracao_por_arquivo_mais_antigo():
    client = _make_client()
    mid = _criar(client).json()["id"]
    _desbloquear(client)
    newer = {"arquivo": ("nova.xlsx", _spreadsheet_bytes(period=datetime(2026, 7, 21)), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    older = {"arquivo": ("antiga.xlsx", _spreadsheet_bytes(period=datetime(2026, 7, 14)), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    assert client.post(
        f"/api/cc/mestre/{mid}/apuracao", files=newer,
        data={"rca_code": "607", "confirmar": "true"},
    ).status_code == 200

    response = client.post(
        f"/api/cc/mestre/{mid}/apuracao", files=older,
        data={"confirmar": "true"},
    )

    assert response.status_code == 409
    assert "mais recente" in response.json()["detail"]
