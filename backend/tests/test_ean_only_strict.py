import os
import sys
import tempfile

import pandas as pd
from openpyxl import Workbook

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from services.excel_processor import (
    _append_cotacao_items_from_dataframe,
    gerar_excel_resultado,
    ler_cotacao,
    ler_tabela_mestre,
)
from services.matching_engine import processar_cotacao


def _xlsx(rows):
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    wb.close()
    tmp.close()
    return tmp.name


def _xlsx_sheets(sheets):
    wb = Workbook()
    for idx, (sheet_name, rows) in enumerate(sheets.items()):
        ws = wb.active if idx == 0 else wb.create_sheet()
        ws.title = sheet_name
        for row in rows:
            ws.append(row)
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    wb.close()
    tmp.close()
    return tmp.name


def test_modo_ean_nao_usa_descricao_quando_ean_nao_bate():
    itens = [{"linha": 2, "ean": "", "nome": "ARROZ CAMIL 5KG"}]
    precos = {"7891234567890": 25.9}
    precos_nome = [{"norm": "ARROZ CAMIL 5KG", "ord": "5KG ARROZ CAMIL", "preco": 25.9, "orig": "ARROZ CAMIL 5KG"}]

    resultado = processar_cotacao(itens, precos, precos_nome, modo="EAN")

    assert resultado == [{"linha": 2, "preco": None, "tipo": None}]


def test_coluna_codigo_na_cotacao_nao_e_tratada_como_ean():
    path = _xlsx([
        ["PRODUTO", "CODIGO", "PRECO"],
        ["ARROZ CAMIL 5KG", "12345678", ""],
    ])
    try:
        itens, _ = ler_cotacao(path)
    finally:
        os.unlink(path)

    assert itens[0]["nome"] == "ARROZ CAMIL 5KG"
    assert itens[0]["ean"] == ""


def test_cotacao_codigo_nome_preco_reconhece_ean_e_nome():
    path = _xlsx([
        ["CÓDIGO", "NOME", "PREÇO"],
        ["7891000379585", "ACHOC. NESCAU 200G", ""],
        ["7891000412855", "ACHOC. PO NESCAU LT 350G", ""],
    ])
    try:
        itens, _ = ler_cotacao(path)
    finally:
        os.unlink(path)

    assert itens[0]["ean"] == "7891000379585"
    assert itens[0]["nome"] == "ACHOC. NESCAU 200G"
    assert itens[0]["col_preco"] == 2


def test_cotacao_prefere_gtin_produto_a_gtin_caixa_e_preenche_preco():
    path = _xlsx([
        ["UNILEVER PERFUMARIA", None, None, None, None],
        ["Referência", "GTIN/PLU Caixa", "GTIN/PLU", "Descrição", "preco"],
        [116183, 27891150071807, 7891150071803, "ANT-SEP BUC CLOSE-UP ICE S/ALCOOL LEV500PG350ML", None],
    ])
    output = None
    try:
        itens, header_row = ler_cotacao(path)

        assert header_row == 2
        assert itens[0]["ean"] == "7891150071803"
        assert itens[0]["nome"] == "ANT-SEP BUC CLOSE-UP ICE S/ALCOOL LEV500PG350ML"
        assert itens[0]["col_preco"] == 4

        output = gerar_excel_resultado(
            path,
            itens,
            [{"linha": 3, "preco": 12.34, "tipo": "EAN"}],
        )

        from openpyxl import load_workbook

        wb = load_workbook(output, data_only=True)
        ws = wb.active
        try:
            assert ws.cell(2, 5).value == "preco"
            assert ws.cell(3, 5).value == 12.34
        finally:
            wb.close()
    finally:
        os.unlink(path)
        if output:
            os.unlink(output)


def test_cotacao_extrai_ean_do_inicio_da_descricao_sem_coluna_ean():
    path = _xlsx([
        ["COTACAO ENVIO", "preco"],
        [None, None],
        ["7891200234257 – SUPERBONDER 5G + 1G GRATIS", None],
        ["7891065000141 -CADEADO PADO 20MM", None],
        ["070330709485 – APARELHO BIC COMFORT 2 SENSITIVE C/2", None],
        ["RAYOVAC PILHA AMARELINHA AAA C/4UN BILISTER", None],
    ])
    output = None
    try:
        itens, header_row = ler_cotacao(path)

        assert header_row == 1
        assert itens[0]["ean"] == "7891200234257"
        assert itens[0]["nome"] == "SUPERBONDER 5G + 1G GRATIS"
        assert itens[0]["col_preco"] == 1
        assert itens[1]["ean"] == "7891065000141"
        assert itens[1]["nome"] == "CADEADO PADO 20MM"
        assert itens[2]["ean"] == "070330709485"
        assert itens[2]["nome"] == "APARELHO BIC COMFORT 2 SENSITIVE C/2"
        assert itens[3]["ean"] == ""
        assert itens[3]["nome"] == "RAYOVAC PILHA AMARELINHA AAA C/4UN BILISTER"

        output = gerar_excel_resultado(
            path,
            itens[:1],
            [{"linha": 3, "preco": 9.99, "tipo": "EAN"}],
        )

        from openpyxl import load_workbook

        wb = load_workbook(output, data_only=True)
        ws = wb.active
        try:
            assert ws.cell(1, 2).value == "preco"
            assert ws.cell(3, 2).value == 9.99
        finally:
            wb.close()
    finally:
        os.unlink(path)
        if output:
            os.unlink(output)


def test_cotacao_nao_substitui_ean_quando_coluna_ean_existe():
    path = _xlsx([
        ["EAN", "Produto", "preco"],
        ["7891000379585", "7899999999999 – ACHOC. NESCAU 200G", ""],
    ])
    try:
        itens, _ = ler_cotacao(path)
    finally:
        os.unlink(path)

    assert itens[0]["ean"] == "7891000379585"
    assert itens[0]["nome"] == "7899999999999 – ACHOC. NESCAU 200G"
    assert itens[0]["col_preco"] == 2


def test_pandas_fallback_preserva_linha_real_do_excel():
    df = pd.DataFrame(
        [["7891000379585", "ACHOC. NESCAU 200G", ""]],
        columns=["Código Barras", "Descrição", "Vlr.Unitário"],
    )
    itens = []

    _append_cotacao_items_from_dataframe(
        itens,
        df,
        header_row=5,
        col_nome=1,
        col_ean=0,
        col_preco=2,
    )

    assert itens == [{
        "ean": "7891000379585",
        "nome": "ACHOC. NESCAU 200G",
        "linha": 6,
        "col_preco": 2,
    }]


def test_coluna_cod_produto_nao_e_tratada_como_descricao():
    path = _xlsx([
        ["Cód. Produto", "Ean", "Descrição", "Emb."],
        [68009, "7896369615077", "ABACAXI CALDA MARIZA LT 400G", 12],
    ])
    try:
        itens, _ = ler_cotacao(path)
    finally:
        os.unlink(path)

    assert itens[0]["nome"] == "ABACAXI CALDA MARIZA LT 400G"
    assert itens[0]["ean"] == "7896369615077"


def test_coluna_codigo_na_tabela_mestre_nao_entra_no_dict_de_ean():
    path = _xlsx([
        ["PRODUTO", "CODIGO", "28"],
        ["ARROZ CAMIL 5KG", "12345678", 25.9],
    ])
    try:
        precos, precos_nome = ler_tabela_mestre(path, prazo=28)
    finally:
        os.unlink(path)

    assert "12345678" not in precos
    assert precos_nome[0]["preco"] == 25.9


def test_tabela_mestre_codigo_nome_preco_reconhece_ean_por_valores():
    path = _xlsx([
        ["CÓDIGO", "NOME", "PREÇO"],
        ["7891000379585", "ACHOC. NESCAU 200G", 8.75],
        ["7891000412855", "ACHOC. PO NESCAU LT 350G", 12.4],
    ])
    try:
        precos, precos_nome = ler_tabela_mestre(path)
    finally:
        os.unlink(path)

    assert precos["7891000379585"] == 8.75
    assert precos_nome[0]["orig"] == "ACHOC. NESCAU 200G"


def test_tabela_mestre_sem_prazo_usa_preco_unitario_e_nao_total():
    path = _xlsx([
        ["Dados da Empresa e do Cliente", None, None, None, None, None, None, None, None, None],
        ["Empresa: TESTE", None, None, None, None, None, None, None, None, None],
        ["Elementos que compõem o Pedido", None, None, None, None, None, None, None, None, None],
        ["Código", "Produto", "EAN", "Emb.", "Qtde", None, "Qtd.\nUN", "R$\nUnit.", "R$\nEmb.", "R$\nTotal"],
        ["3511-18", "ABS ALWAYS BASICO C/8 C/ABA PQ SC", "7500435127226", "CX-18", 1, None, 18, "3.1 21", "56.1 80", "56.1 80"],
        ["Total Venda:", None, None, None, None, None, None, None, None, None],
        ["3531-16", "ABS INTIMUS NOTURNO C/30 SECO", "7896007550906", "CX-16", 1, None, 16, "1 6.1 60", "258.56", "258.56"],
    ])
    try:
        precos, precos_nome = ler_tabela_mestre(path, prazo=28)
    finally:
        os.unlink(path)

    assert precos["7500435127226"] == 3.121
    assert precos["7896007550906"] == 16.160
    assert [item["orig"] for item in precos_nome] == [
        "ABS ALWAYS BASICO C/8 C/ABA PQ SC",
        "ABS INTIMUS NOTURNO C/30 SECO",
    ]


def test_resultado_nao_sobrescreve_coluna_embalagem_quando_nao_ha_preco():
    path = _xlsx([
        ["Cód. Produto", "Ean", "Descrição", "Emb."],
        [68009, "7896369615077", "ABACAXI CALDA MARIZA LT 400G", 12],
    ])
    output = None
    try:
        itens, _ = ler_cotacao(path)
        output = gerar_excel_resultado(
            path,
            itens,
            [{"linha": 2, "preco": 14.87, "tipo": "EAN"}],
        )

        from openpyxl import load_workbook

        wb = load_workbook(output, data_only=True)
        ws = wb.active
        try:
            assert ws.cell(1, 4).value == "Emb."
            assert ws.cell(2, 4).value == 12
            assert ws.cell(1, 5).value == "PRECO"
            assert ws.cell(2, 5).value == 14.87
        finally:
            wb.close()
    finally:
        os.unlink(path)
        if output:
            os.unlink(output)


def test_cotacao_multiplas_abas_preenche_cada_loja_na_aba_correta():
    path = _xlsx_sheets({
        "MATRIZ": [
            ["FALTAS", "EAN", "PREÇO", "OBS"],
            ["ARROZ TESTE 5KG", "7891234567890", None, None],
        ],
        "PRAÇA": [
            ["FALTAS", "EAN", "PREÇO", "OBS"],
            ["FEIJAO TESTE 1KG", "7891234567891", None, None],
        ],
        "RESUMO": [
            ["TOTAL", "OBS"],
            [2, "nao deve virar produto"],
        ],
    })
    output = None
    try:
        itens, header_row = ler_cotacao(path)

        assert header_row == 1
        assert [(i["sheet_name"], i["nome"], i["ean"], i["col_preco"]) for i in itens] == [
            ("MATRIZ", "ARROZ TESTE 5KG", "7891234567890", 2),
            ("PRAÇA", "FEIJAO TESTE 1KG", "7891234567891", 2),
        ]

        output = gerar_excel_resultado(
            path,
            itens,
            [
                {"linha": 2, "preco": 21.5, "tipo": "EAN"},
                {"linha": 2, "preco": 7.25, "tipo": "EAN"},
            ],
        )

        from openpyxl import load_workbook

        wb = load_workbook(output, data_only=True)
        try:
            assert wb["MATRIZ"].cell(2, 3).value == 21.5
            assert wb["PRAÇA"].cell(2, 3).value == 7.25
            assert wb["RESUMO"].cell(1, 3).value in (None, "")
            assert wb["RESUMO"].cell(2, 3).value in (None, "")
        finally:
            wb.close()
    finally:
        os.unlink(path)
        if output:
            os.unlink(output)


def test_resultado_preenche_coluna_custo_sem_criar_preco_extra():
    path = _xlsx([
        ["Cód. Produto", "Ean", "Descrição", "Emb.", "Vlr. Custo"],
        [68009, "7896369615077", "ABACAXI CALDA MARIZA LT 400G", 12, 99.99],
    ])
    output = None
    try:
        itens, _ = ler_cotacao(path)
        output = gerar_excel_resultado(
            path,
            itens,
            [{"linha": 2, "preco": 14.87, "tipo": "EAN"}],
        )

        from openpyxl import load_workbook

        wb = load_workbook(output, data_only=True)
        ws = wb.active
        try:
            # Mantém a escrita na coluna de custo/preço existente (coluna 5)
            assert ws.cell(1, 5).value == "Vlr. Custo"
            assert ws.cell(2, 5).value == 14.87
            # Não deve criar coluna PRECO na frente
            assert ws.cell(1, 6).value in (None, "")
            assert ws.cell(2, 6).value in (None, "")
        finally:
            wb.close()
    finally:
        os.unlink(path)
        if output:
            os.unlink(output)


def test_resultado_ignora_vlr_min_fatur_e_preenche_vlr_unitario():
    path = _xlsx([
        ["Data da Entrega:", None, None, None, None, None, None],
        ["Cond.Pagto:", None, 0, None, None, None, None],
        ["Vlr.Min.Fatur:", None, 0, None, None, None, None],
        [None, None, None, None, None, None, None],
        ["Seq", "Cód.PLU", "Código Barras", "Descrição", "Qt.Emb", "Vlr.Unitário", "%Icms"],
        ["1", "45251-9", "789600754187-4", "ABS INTIMUS INTERNO SUPER C/8", 0, 0, 0],
    ])
    output = None
    try:
        itens, header_row = ler_cotacao(path)

        assert header_row == 5
        assert itens[0]["linha"] == 6
        assert itens[0]["col_preco"] == 5

        output = gerar_excel_resultado(
            path,
            itens,
            [{"linha": 6, "preco": 14.87, "tipo": "EAN"}],
        )

        from openpyxl import load_workbook

        wb = load_workbook(output, data_only=True)
        ws = wb.active
        try:
            assert ws.cell(5, 1).value == "Seq"
            assert ws.cell(6, 1).value == "1"
            assert ws.cell(5, 6).value == "Vlr.Unitário"
            assert ws.cell(6, 6).value == 14.87
        finally:
            wb.close()
    finally:
        os.unlink(path)
        if output:
            os.unlink(output)


def test_cotacao_nao_preenche_quantidade_unitaria_quando_existe_valor_unitario():
    path = _xlsx([
        [
            "Codigo de Barras",
            "Codigo do Produto",
            "Codigo Produto do Fornecedor",
            "Descricao",
            "Complemento",
            "Marca",
            "Departamento",
            "Disponível para venda",
            "Possui Substituição Tributária",
            "Quantidade Unitaria",
            "Valor Unitario",
            "Valor Embalagem",
            "Quantidade por Embalagem",
        ],
        [
            "7896273100881",
            "123",
            None,
            "Aguardente Pirassununga 21 900ml",
            None,
            None,
            None,
            "Sim",
            "Nao",
            None,
            None,
            None,
            1,
        ],
    ])
    output = None
    try:
        itens, header_row = ler_cotacao(path)

        assert header_row == 1
        assert itens[0]["ean"] == "7896273100881"
        assert itens[0]["nome"] == "Aguardente Pirassununga 21 900ml"
        assert itens[0]["col_preco"] == 10

        output = gerar_excel_resultado(
            path,
            itens,
            [{"linha": 2, "preco": 8.63, "tipo": "EAN"}],
        )

        from openpyxl import load_workbook

        wb = load_workbook(output, data_only=True)
        ws = wb.active
        try:
            assert ws.cell(1, 10).value == "Quantidade Unitaria"
            assert ws.cell(2, 10).value in (None, "")
            assert ws.cell(1, 11).value == "Valor Unitario"
            assert ws.cell(2, 11).value == 8.63
        finally:
            wb.close()
    finally:
        os.unlink(path)
        if output:
            os.unlink(output)


def test_cotacao_sbroggio_preenche_coluna_preco_e_inferir_ean_sem_cabecalho():
    path = _xlsx([
        [None, None, None, "FECHAMENTO DA COTACAO", None, None, None, None, None, None, None, None, None, None, None, None],
        [None, None, None, "SEGUNDA FEIRA AS 10HORA", None, None, None, None, None, None, None, None, None, None, None, None],
        ["BRUNO", "SBROGGIO SUPERMERCADO LTDA  LOJA 02", None, None, "PLANILHA DE COTACAO LOJA 02", None, None, None, None, None, None, None, None, None, None, None],
        [None, None, None, None, "BETO", "preco", "GABRIEL", "GLALCIO", "IKEDA", "MARCOS", "THIAGO", "CIDINHA", "LUIZ", "JUNIOR", "ATACAD", "preco"],
        ["DATA", None, None, None, None, None, None, None, None, None, None, None, None, None, None, None],
        ["2026-06-01", None, None, "PRODUTOS", "PAULIST", "preco", "SERVIMED", "TOP SERV", "IKEDA", "MERC", "FR DISTRI", "CIDINHA", "VAQUEIRO", "V.NOVA", "ATACAD", "preco"],
        [19.99, "4005900715814", "12 UNIDADES", "ANT NIVEA DEO MILK 150ML", None, None, None, None, None, None, None, None, None, None, None, None],
        [14.99, "7896004400341", "1 CX", "FLOCOCO SOCOCO 100 G.", None, None, None, None, None, None, None, None, None, None, None, None],
    ])
    output = None
    try:
        itens, header_row = ler_cotacao(path)

        assert header_row == 6
        assert itens[0]["ean"] == "4005900715814"
        assert itens[0]["nome"] == "ANT NIVEA DEO MILK 150ML"
        assert itens[0]["col_preco"] == 5

        output = gerar_excel_resultado(
            path,
            itens[:1],
            [{"linha": 7, "preco": 14.87, "tipo": "EAN"}],
        )

        from openpyxl import load_workbook

        wb = load_workbook(output, data_only=True)
        ws = wb.active
        try:
            assert ws.cell(6, 6).value == "preco"
            assert ws.cell(7, 6).value == 14.87
            assert ws.cell(6, 17).value in (None, "")
            assert ws.cell(7, 17).value in (None, "")
        finally:
            wb.close()
    finally:
        os.unlink(path)
        if output:
            os.unlink(output)
