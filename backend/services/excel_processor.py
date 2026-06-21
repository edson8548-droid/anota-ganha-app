"""
Processador de Excel — leitura de tabelas mestre e cotações, geração de resultado.
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import tempfile
import os
import io
import zipfile
import re as _re
import unicodedata

from .matching_engine import limpar_ean, normalizar_nome, ordenar_palavras, processar_cotacao_com_ia


def _normalizar_cabecalho(valor) -> str:
    c = str(valor or "").upper().strip()
    c = unicodedata.normalize("NFKD", c).encode("ASCII", "ignore").decode("ascii")
    c = _re.sub(r"[^A-Z0-9]+", " ", c)
    return _re.sub(r"\s+", " ", c).strip()


def _cabecalho_sem_nome(valor) -> bool:
    c_norm = _normalizar_cabecalho(valor)
    return not c_norm or c_norm.startswith("UNNAMED ")


def _score_coluna_nome(nome_coluna) -> int:
    c_norm = _normalizar_cabecalho(nome_coluna)
    if not c_norm:
        return 0
    if "PRODUTO" in c_norm and any(k in c_norm for k in ("COD", "CODIGO")):
        return 0
    if c_norm in {"NOME", "NOME PRODUTO", "NOME DO PRODUTO", "DESCRICAO", "DESCRICAO PRODUTO"}:
        return 100
    if any(k in c_norm for k in ("PRODUTO", "DESCRI", "ITEM", "MERCAD")):
        return 80
    return 0


def _score_coluna_ean(nome_coluna) -> int:
    """Pontua colunas prováveis de EAN/GTIN evitando confundir com código interno."""
    c_norm = _normalizar_cabecalho(nome_coluna)

    if not c_norm:
        return 0
    embalagem_penalty = 10 if _re.search(r"\b(CAIXA|CX|EMB|EMBALAGEM|DUN|MASTER|PACK)\b", c_norm) else 0
    if (
        "COD PROD" in c_norm
        or "CODIGO PROD" in c_norm
        or "COD INTERNO" in c_norm
        or "CODIGO INTERNO" in c_norm
        or c_norm in {"COD", "CODIGO"}
    ):
        return 0
    if "EAN" in c_norm or "GTIN" in c_norm:
        return 100 - embalagem_penalty
    if "CODIGO DE BARRAS" in c_norm or "COD BARRAS" in c_norm or "COD BARRA" in c_norm:
        return 95 - embalagem_penalty
    if "COD BARR" in c_norm or "CODBAR" in c_norm or "COD BARRA" in c_norm:
        return 92 - embalagem_penalty
    if "BARRAS" in c_norm or "BARRA" in c_norm or "BARCODE" in c_norm:
        return 90 - embalagem_penalty
    return 0


def _melhor_coluna_ean(colunas):
    melhor_col = None
    melhor_score = 0
    for col in colunas:
        score = _score_coluna_ean(col)
        if score > melhor_score:
            melhor_col = col
            melhor_score = score
    return melhor_col if melhor_score >= 80 else None


def _permite_inferir_ean_por_valores(nome_coluna) -> bool:
    c_norm = _normalizar_cabecalho(nome_coluna)
    if _cabecalho_sem_nome(nome_coluna):
        return True
    if any(k in c_norm for k in ("COD PROD", "CODIGO PROD", "COD INTERNO", "CODIGO INTERNO", "INTERNO")):
        return False
    return (
        c_norm in {"COD", "CODIGO"}
        or "EAN" in c_norm
        or "GTIN" in c_norm
        or "BARRA" in c_norm
        or "BARCODE" in c_norm
    )


def _parece_ean_real(valor) -> bool:
    ean = limpar_ean(valor)
    return 12 <= len(ean) <= 14


def _inferir_coluna_ean_openpyxl(ws, header_row: int, ignorar_cols=None):
    ignorar_cols = {c for c in (ignorar_cols or set()) if c is not None}
    melhor_col = None
    melhor_score = 0

    for col_idx in range(ws.max_column):
        if col_idx in ignorar_cols:
            continue
        header_value = ws.cell(row=header_row, column=col_idx + 1).value
        if not _permite_inferir_ean_por_valores(header_value):
            continue

        total = 0
        hits = 0
        for row_idx in range(header_row + 1, min(ws.max_row, header_row + 50) + 1):
            value = ws.cell(row=row_idx, column=col_idx + 1).value
            if value is None or str(value).strip() == "":
                continue
            total += 1
            if _parece_ean_real(value):
                hits += 1

        if hits and total:
            score = hits / total
            min_hits = 2 if _cabecalho_sem_nome(header_value) else 1
            if hits >= min_hits and score >= 0.8 and score > melhor_score:
                melhor_col = col_idx
                melhor_score = score

    return melhor_col


def _inferir_coluna_ean_dataframe(df, ignorar_cols=None):
    ignorar_cols = {c for c in (ignorar_cols or set()) if c is not None}
    melhor_col = None
    melhor_score = 0

    for col_idx, col in enumerate(df.columns):
        if col_idx in ignorar_cols:
            continue
        if not _permite_inferir_ean_por_valores(col):
            continue

        values = [v for v in df.iloc[:50, col_idx].tolist() if v is not None and str(v).strip() not in ("", "nan")]
        if not values:
            continue
        hits = sum(1 for v in values if _parece_ean_real(v))
        score = hits / len(values)
        min_hits = 2 if _cabecalho_sem_nome(col) else 1
        if hits >= min_hits and score >= 0.8 and score > melhor_score:
            melhor_col = col_idx
            melhor_score = score

    return melhor_col


def _score_coluna_preco(nome_coluna, prazo=None) -> int:
    """Pontua colunas de preco evitando total, embalagem e quantidade."""
    c_norm = _normalizar_cabecalho(nome_coluna)

    if not c_norm:
        return 0

    if any(k in c_norm for k in ("TOTAL", "EMB", "CAIXA", "CX", "QTD", "QTDE", "QUANT")):
        return 0

    if prazo:
        nums = _re.findall(r"\b(\d+)\b", c_norm)
        if str(prazo) in nums:
            return 100

    if "UNIT" in c_norm or "UNITARIO" in c_norm or "UNITARIA" in c_norm:
        return 95
    if "PRECO UNIT" in c_norm or "VALOR UNIT" in c_norm:
        return 95
    if c_norm in {"R UNIT", "RS UNIT", "R UNITARIO", "RS UNITARIO"}:
        return 95
    if c_norm in {"PRECO", "PRE O", "VALOR", "R", "RS"}:
        return 60
    if "PRECO" in c_norm or "VALOR" in c_norm:
        return 55
    return 0


def _melhor_coluna_preco(colunas, prazo=None):
    melhor_col = None
    melhor_score = 0
    for col in colunas:
        score = _score_coluna_preco(col, prazo=prazo)
        if score > melhor_score:
            melhor_col = col
            melhor_score = score
    return melhor_col if melhor_score >= 55 else None


def _score_coluna_fracionamento(nome_coluna) -> int:
    c_norm = _normalizar_cabecalho(nome_coluna)
    if not c_norm:
        return 0

    if any(k in c_norm for k in ("PRECO", "VALOR", "TOTAL", "R ", "RS ", "EAN", "GTIN", "BARRA")):
        return 0
    if any(k in c_norm for k in ("COD PROD", "CODIGO PROD", "COD INTERNO", "CODIGO INTERNO")):
        return 0

    if "FRACION" in c_norm:
        return 100
    if any(k in c_norm for k in ("QTD CAIXA", "QTDE CAIXA", "QUANTIDADE CAIXA", "QUANT CAIXA")):
        return 98
    if any(k in c_norm for k in ("QTD CX", "QTDE CX", "QUANT CX", "QTE CX")):
        return 96
    if any(k in c_norm for k in ("UN CAIXA", "UND CAIXA", "UN POR CAIXA", "UND POR CAIXA")):
        return 94
    if "MULTIPLO" in c_norm:
        return 90
    if c_norm in {"CX", "CAIXA", "QTD", "QTDE", "QUANTIDADE", "EMB", "EMBALAGEM"}:
        return 84
    if any(k in c_norm for k in ("QTD EMB", "QTDE EMB", "QUANT EMB", "QTD EMBALAGEM")):
        return 88
    return 0


def _melhor_coluna_fracionamento(colunas, ignorar_cols=None):
    ignorar_cols = {c for c in (ignorar_cols or set()) if c is not None}
    melhor_col = None
    melhor_score = 0
    for col in colunas:
        if col in ignorar_cols:
            continue
        score = _score_coluna_fracionamento(col)
        if score > melhor_score:
            melhor_col = col
            melhor_score = score
    return melhor_col if melhor_score >= 84 else None


def _parse_fracionamento(valor):
    if valor is None:
        return None
    try:
        if pd.isna(valor):
            return None
    except Exception:
        pass

    raw = str(valor).strip()
    if not raw or raw.upper() in {"NONE", "NAN", "NULL"}:
        return None

    match = _re.search(r"\d+(?:[.,]\d+)?", raw)
    if not match:
        return None

    try:
        numero = float(match.group(0).replace(",", "."))
    except ValueError:
        return None

    if not numero or numero <= 0 or numero > 9999:
        return None

    if abs(numero - round(numero)) < 0.0001:
        return str(int(round(numero)))
    return (f"{numero:.3f}".rstrip("0").rstrip(".")).replace(".", ",")


def _xlsx_safe_bytes(caminho_arquivo):
    """
    Retorna BytesIO do xlsx com styles.xml corrigido.
    Alguns arquivos gerados por sistemas ERP têm indent > 255 que quebra o openpyxl.
    """
    try:
        buf = io.BytesIO()
        with zipfile.ZipFile(caminho_arquivo) as zin, \
             zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == 'xl/styles.xml':
                    try:
                        styles = data.decode('utf-8', errors='replace')
                        styles = _re.sub(
                            r'indent="(\d+)"',
                            lambda m: 'indent="255"' if int(m.group(1)) > 255 else m.group(0),
                            styles,
                        )
                        data = styles.encode('utf-8')
                    except Exception:
                        pass
                zout.writestr(item, data)
        buf.seek(0)
        return buf
    except Exception:
        # Não é um zip válido ou outro erro — devolve o arquivo original como BytesIO
        with open(caminho_arquivo, 'rb') as f:
            return io.BytesIO(f.read())

PREENCHIMENTO_IA = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
MAX_BLANK_ROWS_AFTER_COTACAO_ITEMS = 200
_PRECO_COTACAO_KW = (
    "PREÇO", "PRECO", "VALOR UNIT", "VALOR UNI", "VALOR", "R$",
    "PRECO UNIT", "PREÇO UNIT", "UNIT", "CUSTO", "VLR",
    "VLR. CUSTO", "VLR CUSTO",
)


def _extrair_ean_prefixo_nome(nome_val):
    """Extrai EAN/GTIN no início da descrição sem pegar números do meio do texto."""
    texto = str(nome_val or "").strip()
    if not texto:
        return "", texto

    match = _re.match(
        r"^\s*['\"]?((?:\d[\s.-]?){8,14})\s*(?:[-–—:|]\s*)+(.+?)\s*$",
        texto,
    )
    if not match:
        return "", texto

    ean = _re.sub(r"\D", "", match.group(1))
    nome_sem_ean = str(match.group(2) or "").strip()
    if not (8 <= len(ean) <= 14) or not nome_sem_ean:
        return "", texto

    return ean, nome_sem_ean


def _cotacao_item(ean_val, nome_val, linha, col_preco, sheet_name=None):
    nome_texto = str(nome_val).strip()
    ean_texto = str(ean_val) if ean_val and str(ean_val) != "nan" else ""

    if not ean_texto:
        ean_prefixo, nome_sem_ean = _extrair_ean_prefixo_nome(nome_texto)
        if ean_prefixo:
            ean_texto = ean_prefixo
            nome_texto = nome_sem_ean

    item = {
        "ean": ean_texto,
        "nome": nome_texto,
        "linha": linha,
        "col_preco": col_preco,
    }
    if sheet_name:
        item["sheet_name"] = sheet_name
    return item


def _append_cotacao_items_from_dataframe(itens, df, header_row, col_nome, col_ean, col_preco):
    """Adiciona itens lidos por pandas preservando a linha real do Excel."""
    for idx, row in df.iterrows():
        nome_val = row.iloc[col_nome]
        ean_val = row.iloc[col_ean] if col_ean is not None and col_ean < len(row) else ""
        if nome_val and str(nome_val).strip() and str(nome_val).strip() != "nan":
            itens.append(_cotacao_item(
                ean_val,
                nome_val,
                header_row + idx + 1,
                col_preco,
            ))


def _match_cotacao_nome(val):
    return _score_coluna_nome(val) >= 80


def _match_cotacao_preco(val):
    if _score_coluna_preco(val) >= 55:
        return True
    c_norm = _normalizar_cabecalho(val)
    return bool(_re.search(r"\bVLR\b.*\bCUSTO\b", c_norm))


def _ler_cotacao_worksheet(ws, sheet_name=None, exigir_indicio_cotacao=False):
    """Lê itens de uma aba de cotação preservando linha, coluna de preço e aba."""
    col_ean = None
    col_nome = None
    col_preco = None
    found_header = None
    best_header = None
    best_score = -1
    fallback_col_ean = None
    fallback_col_nome = None
    fallback_col_preco = None
    fallback_header = None
    fallback_col_ean_score = 0

    for row_idx in range(1, min(16, ws.max_row + 1)):
        row_col_ean = None
        row_col_nome = None
        row_col_preco = None
        row_col_ean_score = 0
        for cell in ws[row_idx]:
            val = str(cell.value).upper().strip() if cell.value else ""
            if not val:
                continue

            ean_score = _score_coluna_ean(val)
            if ean_score >= 80:
                if ean_score > row_col_ean_score:
                    row_col_ean = cell.column - 1
                    row_col_ean_score = ean_score
                if ean_score > fallback_col_ean_score:
                    fallback_col_ean = cell.column - 1
                    fallback_col_ean_score = ean_score
                    fallback_header = fallback_header or row_idx
                continue

            if row_col_nome is None and _match_cotacao_nome(val):
                row_col_nome = cell.column - 1
            elif row_col_preco is None and _match_cotacao_preco(val):
                row_col_preco = cell.column - 1

            if fallback_col_nome is None and _match_cotacao_nome(val):
                fallback_col_nome = cell.column - 1
                fallback_header = row_idx
            elif fallback_col_preco is None and _match_cotacao_preco(val):
                fallback_col_preco = cell.column - 1
                fallback_header = fallback_header or row_idx

        if row_col_nome is not None:
            score = 4
            if row_col_ean is not None:
                score += 3
            if row_col_preco is not None:
                score += 2
            if score > best_score:
                best_score = score
                best_header = (row_idx, row_col_nome, row_col_ean, row_col_preco)

    if best_header:
        found_header, col_nome, col_ean, col_preco = best_header
    else:
        found_header = fallback_header
        col_nome = fallback_col_nome
        col_ean = fallback_col_ean
        col_preco = fallback_col_preco

    if exigir_indicio_cotacao and found_header is None:
        return [], 1

    header_row = found_header or 1
    if col_ean is None:
        col_ean = _inferir_coluna_ean_openpyxl(
            ws,
            header_row,
            ignorar_cols={col_nome, col_preco},
        )
    if col_nome is None:
        col_nome = 0

    itens = []
    blank_rows_after_items = 0
    scan_col_limit = max(
        20,
        (col_nome or 0) + 1,
        (col_ean or 0) + 1,
        (col_preco or 0) + 1,
    )

    for row_idx in range(header_row + 1, ws.max_row + 1):
        row = ws[row_idx]
        ean_val = row[col_ean].value if col_ean is not None and col_ean < len(row) else None
        nome_val = row[col_nome].value if col_nome < len(row) else None
        if nome_val and str(nome_val).strip() and str(nome_val).strip().upper() not in ("NONE", "NAN"):
            itens.append(_cotacao_item(ean_val, nome_val, row_idx, col_preco, sheet_name=sheet_name))
            blank_rows_after_items = 0
        elif itens:
            row_has_value = any(
                cell.value is not None and str(cell.value).strip() != ""
                for cell in row[:scan_col_limit]
            )
            if row_has_value:
                blank_rows_after_items = 0
            else:
                blank_rows_after_items += 1
                if blank_rows_after_items >= MAX_BLANK_ROWS_AFTER_COTACAO_ITEMS:
                    break

    return itens, header_row


def detectar_prazos_disponiveis(caminho_arquivo) -> list:
    """Detecta quais colunas de prazo (7, 14, 21, 28 dias) existem no Excel."""
    for header in range(0, 10):
        try:
            df = pd.read_excel(caminho_arquivo, header=header, nrows=0)
            encontrados = []
            for prazo in [7, 14, 21, 28]:
                for col in df.columns:
                    nums = _re.findall(r'\b(\d+)\b', str(col))
                    if str(prazo) in nums:
                        encontrados.append(prazo)
                        break
            if encontrados:
                return sorted(encontrados)
        except Exception:
            continue
    return [28]


def ler_tabela_mestre(caminho_arquivo, header_row=None, col_nome=0, col_ean=1, prazo=28, incluir_meta=False):
    """
    Lê Excel de tabela de preços mestre. Auto-detecta linha de cabeçalho e coluna do prazo.
    Retorna: (precos_dict, precos_nome_lista) ou, com incluir_meta=True,
    (precos_dict, precos_nome_lista, meta_por_ean).
    """
    buf = _xlsx_safe_bytes(caminho_arquivo)

    df_final = None
    col_nome_final = None
    col_ean_final = None
    col_preco_final = None
    col_fracionamento_final = None

    for hdr in range(0, 12):
        try:
            buf.seek(0)
            df = pd.read_excel(buf, header=hdr)
            if df.empty or len(df.columns) < 2:
                continue

            cols = df.columns
            # Coluna de preço: prazo quando existir; senão preço unitário.
            c_preco = _melhor_coluna_preco(cols, prazo=prazo)

            # Coluna de nome
            c_nome = None
            for i, c in enumerate(cols):
                if _score_coluna_nome(c) >= 80:
                    c_nome = cols[i]
                    break

            # Coluna de EAN/GTIN. Prioriza códigos de barras reais e evita COD PRODUTO.
            c_ean = _melhor_coluna_ean(cols)
            if c_ean is None:
                idx_ean = _inferir_coluna_ean_dataframe(
                    df,
                    ignorar_cols={
                        df.columns.get_loc(c_nome) if c_nome is not None else None,
                        df.columns.get_loc(c_preco) if c_preco is not None else None,
                    },
                )
                if idx_ean is not None:
                    c_ean = cols[idx_ean]

            if c_nome is not None and c_preco is not None:
                ignorar_fracionamento = {c_nome, c_ean, c_preco}
                c_fracionamento = _melhor_coluna_fracionamento(cols, ignorar_cols=ignorar_fracionamento)
                df_final = df
                col_nome_final = c_nome
                col_ean_final = c_ean
                col_preco_final = c_preco
                col_fracionamento_final = c_fracionamento
                break
        except Exception:
            continue

    # Fallback: last header tried, use last column as price
    if df_final is None:
        try:
            buf.seek(0)
            df_final = pd.read_excel(buf, header=2)
            col_nome_final = df_final.columns[0]
            col_ean_final = _melhor_coluna_ean(df_final.columns)
            col_preco_final = df_final.columns[-1]
            col_fracionamento_final = _melhor_coluna_fracionamento(
                df_final.columns,
                ignorar_cols={col_nome_final, col_ean_final, col_preco_final},
            )
        except Exception:
            return ({}, [], {}) if incluir_meta else ({}, [])

    precos = {}
    precos_nome_lista = []
    meta_por_ean = {}

    for _, row in df_final.iterrows():
        nome_bruto = str(row[col_nome_final]) if col_nome_final is not None else ""
        if not nome_bruto or nome_bruto.strip().upper() in ("NONE", "NAN", ""):
            continue

        ean_raw = row[col_ean_final] if col_ean_final is not None else None
        ean = limpar_ean(ean_raw)
        fracionamento = _parse_fracionamento(row[col_fracionamento_final]) if col_fracionamento_final is not None else None

        try:
            preco = float(str(row[col_preco_final]).replace(",", ".").replace("R$", "").replace(" ", "").strip())
            if pd.isna(preco) or preco <= 0:
                continue
        except (ValueError, TypeError):
            continue

        if ean:
            precos[ean] = preco
            if fracionamento:
                meta_por_ean[ean] = {"fracionamento": fracionamento}
        nome_norm = normalizar_nome(nome_bruto)
        item_nome = {
            'norm': nome_norm,
            'ord': ordenar_palavras(nome_norm),
            'preco': preco,
            'orig': nome_bruto,
        }
        if fracionamento:
            item_nome['fracionamento'] = fracionamento
        precos_nome_lista.append(item_nome)

    if incluir_meta:
        return precos, precos_nome_lista, meta_por_ean
    return precos, precos_nome_lista


def ler_cotacao(caminho_arquivo):
    """
    Lê Excel de cotação enviado pelo RCA.
    Corrige automaticamente xlsx com XML inválido (indent > 255).
    Usa detecção parcial de cabeçalhos para tolerar variações de nome.
    Retorna: (itens, header_row)
    """
    itens = []
    header_row = 1

    # --- Tentativa 1: openpyxl com correção de stylesheet ---
    try:
        buf = _xlsx_safe_bytes(caminho_arquivo)
        wb = openpyxl.load_workbook(buf, data_only=True)
        varias_abas = len(wb.worksheets) > 1
        first_header_row = None
        for ws in wb.worksheets:
            sheet_items, sheet_header_row = _ler_cotacao_worksheet(
                ws,
                sheet_name=ws.title if varias_abas else None,
                exigir_indicio_cotacao=varias_abas,
            )
            if sheet_items:
                itens.extend(sheet_items)
                if first_header_row is None:
                    first_header_row = sheet_header_row

        if first_header_row is not None:
            header_row = first_header_row
        wb.close()
        if itens:
            return itens, header_row
    except Exception:
        pass

    # --- Tentativa 2: pandas com correção de stylesheet ---
    try:
        buf = _xlsx_safe_bytes(caminho_arquivo)
        for hdr in range(0, 8):
            try:
                df = pd.read_excel(buf, header=hdr)
                buf.seek(0)
                cols_upper = [str(c).upper().strip() for c in df.columns]

                col_nome = col_ean = col_preco = None
                col_ean_score = 0
                for i, c in enumerate(cols_upper):
                    ean_score = _score_coluna_ean(c)
                    if ean_score >= 80:
                        if ean_score > col_ean_score:
                            col_ean = i
                            col_ean_score = ean_score
                        continue
                    if col_nome is None and _match_cotacao_nome(c):
                        col_nome = i
                    elif col_preco is None and _match_cotacao_preco(c):
                        col_preco = i

                if col_nome is None:
                    continue
                if col_ean is None:
                    col_ean = _inferir_coluna_ean_dataframe(df, ignorar_cols={col_nome, col_preco})

                header_row = hdr + 1

                _append_cotacao_items_from_dataframe(
                    itens,
                    df,
                    header_row,
                    col_nome,
                    col_ean,
                    col_preco,
                )
                if itens:
                    break
            except Exception:
                buf.seek(0)
                continue
    except Exception:
        pass

    return itens, header_row


def _cell_is_empty(cell) -> bool:
    value = cell.value
    if value is None:
        return True
    return str(value).strip() == ""


def _first_empty_price_cell(ws, row_idx: int, preferred_zero_based, fallback_start_col: int):
    """
    Retorna uma célula segura para gravar preço sem sobrescrever dados existentes.

    Se a cotação tiver uma coluna de preço detectada e a célula da linha estiver vazia,
    usa essa coluna. Caso contrário, grava na primeira coluna vazia à direita das
    colunas originais da planilha.
    """
    if preferred_zero_based is not None:
        preferred_col = preferred_zero_based + 1
        # Sempre prefere a coluna detectada de preço/custo para evitar criar
        # coluna nova "PRECO" quando a planilha já possui a coluna alvo.
        return ws.cell(row=row_idx, column=preferred_col)

    start_col = fallback_start_col

    col = max(start_col, 1)
    while not _cell_is_empty(ws.cell(row=row_idx, column=col)):
        col += 1
    return ws.cell(row=row_idx, column=col)


def _write_resultados_to_worksheet(ws, itens, resultados):
    fallback_start_col = ws.max_column + 1
    header_row_idx = max(1, min((item.get("linha", 2) for item in itens), default=2) - 1)

    for item, res in zip(itens, resultados):
        if res["preco"] is not None:
            cell = _first_empty_price_cell(
                ws,
                row_idx=item["linha"],
                preferred_zero_based=item.get("col_preco"),
                fallback_start_col=fallback_start_col,
            )
            header_cell = ws.cell(row=header_row_idx, column=cell.column)
            if _cell_is_empty(header_cell):
                header_cell.value = "PRECO"
            cell.value = res["preco"]

            if res["tipo"] and "IA" in res["tipo"]:
                cell.fill = PREENCHIMENTO_IA


def _xlrd_cell_value(cell, datemode):
    import xlrd

    if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
        return None
    if cell.ctype == xlrd.XL_CELL_DATE:
        try:
            return xlrd.xldate.xldate_as_datetime(cell.value, datemode)
        except Exception:
            return cell.value
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype == xlrd.XL_CELL_ERROR:
        return xlrd.biffh.error_text_from_code.get(cell.value)
    return cell.value


def _safe_sheet_title(title, index, used_titles):
    base = _re.sub(r"[\[\]\:\*\?\/\\]+", " ", str(title or "").strip())
    base = _re.sub(r"\s+", " ", base).strip() or f"Planilha {index}"
    base = base[:31]
    candidate = base
    suffix = 2
    while candidate in used_titles:
        marker = f" {suffix}"
        candidate = f"{base[:31 - len(marker)]}{marker}"
        suffix += 1
    used_titles.add(candidate)
    return candidate


def _workbook_from_xls_values(caminho_original):
    """Converte .xls legado para Workbook preservando coordenadas das celulas."""
    import xlrd

    book = xlrd.open_workbook(caminho_original, formatting_info=False)
    wb = Workbook()
    used_titles = set()

    for sheet_idx in range(book.nsheets):
        sheet = book.sheet_by_index(sheet_idx)
        title = _safe_sheet_title(sheet.name, sheet_idx + 1, used_titles)
        ws = wb.active if sheet_idx == 0 else wb.create_sheet(title)
        ws.title = title

        for row_idx in range(sheet.nrows):
            for col_idx in range(sheet.ncols):
                value = _xlrd_cell_value(sheet.cell(row_idx, col_idx), book.datemode)
                if value is not None:
                    ws.cell(row=row_idx + 1, column=col_idx + 1).value = value

    return wb


def _write_resultados_to_workbook(wb, itens, resultados):
    active_sheet = wb.active.title
    grupos = {}
    ordem = []

    for item, res in zip(itens, resultados):
        sheet_name = item.get("sheet_name") or active_sheet
        if sheet_name not in grupos:
            grupos[sheet_name] = ([], [])
            ordem.append(sheet_name)
        grupos[sheet_name][0].append(item)
        grupos[sheet_name][1].append(res)

    for sheet_name in ordem:
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        sheet_items, sheet_results = grupos[sheet_name]
        _write_resultados_to_worksheet(ws, sheet_items, sheet_results)


def gerar_excel_resultado(caminho_original, itens, resultados):
    """
    Gera Excel preenchido com os precos encontrados.
    Itens matched por IA ficam em amarelo.
    Fallback: recria com pandas se openpyxl nao conseguir ler.
    Retorna caminho do arquivo gerado.
    """
    try:
        buf = _xlsx_safe_bytes(caminho_original)
        wb = openpyxl.load_workbook(buf)
        _write_resultados_to_workbook(wb, itens, resultados)

        output = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(output.name)
        output.close()
        return output.name

    except Exception:
        try:
            wb = _workbook_from_xls_values(caminho_original)
            _write_resultados_to_workbook(wb, itens, resultados)

            output = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            wb.save(output.name)
            output.close()
            return output.name
        except Exception:
            pass

        # Fallback: gerar novo Excel com pandas
        df = pd.read_excel(caminho_original)
        preco_col = None
        for c in df.columns:
            cu = str(c).upper().strip()
            if cu in ("PREÇO", "PRECO", "VALOR", "R$"):
                preco_col = c
                break
        if preco_col is None:
            df["PRECO"] = None
            preco_col = "PRECO"

        for item, res in zip(itens, resultados):
            if res["preco"] is not None:
                row_idx = item["linha"] - 2  # ajuste header
                if 0 <= row_idx < len(df):
                    target_col = preco_col
                    existing = df.at[df.index[row_idx], target_col] if target_col in df.columns else None
                    if existing is not None and str(existing).strip() not in ("", "nan", "None"):
                        base_name = "PRECO"
                        target_col = base_name
                        suffix = 2
                        while target_col in df.columns and str(df.at[df.index[row_idx], target_col]).strip() not in ("", "nan", "None"):
                            target_col = f"{base_name}_{suffix}"
                            suffix += 1
                    if target_col not in df.columns:
                        df[target_col] = None
                    df.at[df.index[row_idx], target_col] = res["preco"]

        output = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        df.to_excel(output.name, index=False)
        output.close()
        return output.name


def processar_arquivo_cotacao(caminho_cotacao, caminho_mestre, prazo=28, modo="completo"):
    """
    Pipeline completo: le tabela + cotacao + matching + gera resultado.
    Retorna (caminho_resultado, stats, itens_sem_match)
    """
    precos_dict, precos_lista = ler_tabela_mestre(caminho_mestre, prazo=prazo)
    itens, header_row = ler_cotacao(caminho_cotacao)

    resultados = processar_cotacao_com_ia(itens, precos_dict, precos_lista, modo=modo)
    caminho_resultado = gerar_excel_resultado(caminho_cotacao, itens, resultados)

    stats = {"ean": 0, "descricao": 0, "ia": 0, "sem_match": 0, "total": len(resultados)}
    sem_match = []
    for item, res in zip(itens, resultados):
        if res["tipo"] is None:
            stats["sem_match"] += 1
            sem_match.append(item["nome"])
        elif res["tipo"] == "EAN":
            stats["ean"] += 1
        elif "IA" in res["tipo"]:
            stats["ia"] += 1
        else:
            stats["descricao"] += 1

    return caminho_resultado, stats, sem_match


def _parse_preco(valor) -> float | None:
    """Converte string/number para float de preço, retorna None se inválido."""
    if valor is None:
        return None
    try:
        v = str(valor).replace(",", ".").replace("R$", "").replace(" ", "").strip()
        f = float(v)
        return f if f > 0 else None
    except (ValueError, TypeError):
        return None


def _parse_ean(valor) -> str:
    return limpar_ean(valor)


def _ler_excel_base(caminho_arquivo):
    """
    Lê Excel base do atacadista. Retorna lista de {"nome", "ean", "preco_base"}.
    Estratégia: tenta pandas em múltiplas linhas de header; openpyxl como fallback.
    """
    import logging
    logger_local = logging.getLogger(__name__)

    NOME_KEYWORDS = ("PRODUTO", "DESCRI", "ITEM", "MERCAD", "NOME DO PROD")
    EAN_KEYWORDS  = ("EAN", "COD.BARRAS", "COD BARRAS", "CODIGO", "BARRAS", "BARRA", "GTIN")
    PRECO_KEYWORDS = ("PREÇO", "PRECO", "VALOR", "R$", "UNIT", "28", "21", "14", " 7 ", "7D", "14D", "21D", "28D")

    def _detect_cols(cols_upper):
        col_nome = col_ean = col_preco = None
        for i, c in enumerate(cols_upper):
            if col_nome is None and any(k in c for k in NOME_KEYWORDS):
                col_nome = i
            if col_ean is None and _score_coluna_ean(c) >= 35:
                col_ean = i
            if col_preco is None and any(k in c for k in PRECO_KEYWORDS):
                col_preco = i
        return col_nome, col_ean, col_preco

    def _rows_from_df(df, col_nome, col_ean, col_preco):
        rows = []
        for _, row in df.iterrows():
            nome = str(row.iloc[col_nome]).strip() if col_nome is not None else ""
            if not nome or nome.upper() in ("NONE", "NAN", ""):
                continue
            ean = _parse_ean(row.iloc[col_ean] if col_ean is not None and col_ean < len(row) else None)
            preco_raw = row.iloc[col_preco] if col_preco is not None and col_preco < len(row) else None
            preco = _parse_preco(preco_raw)
            if preco is None:
                continue
            rows.append({"nome": nome, "ean": ean, "preco_base": preco})
        return rows

    # --- Tentativa 1: pandas com detecção de header ---
    for hdr in range(0, 15):
        try:
            df = pd.read_excel(caminho_arquivo, header=hdr)
            if df.empty or len(df.columns) < 2:
                continue
            cols_upper = [str(c).upper().strip() for c in df.columns]
            col_nome, col_ean, col_preco = _detect_cols(cols_upper)

            if col_nome is None:
                continue  # linha de header não identificada, tenta a próxima

            # Se não achou coluna de preço pelo nome, pega a última coluna numérica
            if col_preco is None:
                for i in range(len(df.columns) - 1, -1, -1):
                    sample = df[df.columns[i]].dropna()
                    numeric_count = sum(1 for v in sample if _parse_preco(v) is not None)
                    if numeric_count >= max(3, len(sample) * 0.3):
                        col_preco = i
                        break

            if col_preco is None:
                col_preco = len(df.columns) - 1

            rows = _rows_from_df(df, col_nome, col_ean, col_preco)
            if rows:
                logger_local.info(f"_ler_excel_base: {len(rows)} produtos (pandas header={hdr})")
                return rows
        except Exception as e:
            logger_local.debug(f"pandas header={hdr} falhou: {e}")
            continue

    # --- Tentativa 2: openpyxl direto (fallback para workbooks protegidos/especiais) ---
    try:
        wb_in = openpyxl.load_workbook(caminho_arquivo, data_only=True)
        ws_in = wb_in.active

        header_row_idx = 1
        col_nome = 0
        col_ean = None
        col_preco = None

        for row_idx in range(1, min(20, ws_in.max_row + 1)):
            for cell in ws_in[row_idx]:
                val = str(cell.value).upper().strip() if cell.value else ""
                if col_nome == 0 and any(k in val for k in NOME_KEYWORDS):
                    col_nome = cell.column - 1
                    header_row_idx = row_idx
                elif col_ean is None and _score_coluna_ean(val) >= 35:
                    col_ean = cell.column - 1
                elif col_preco is None and any(k in val for k in PRECO_KEYWORDS):
                    col_preco = cell.column - 1

        if col_preco is None:
            col_preco = ws_in.max_column - 1

        rows_data = []
        for row_idx in range(header_row_idx + 1, ws_in.max_row + 1):
            row = ws_in[row_idx]
            nome = row[col_nome].value if col_nome < len(row) else None
            ean_raw = row[col_ean].value if col_ean is not None and col_ean < len(row) else None
            preco_raw = row[col_preco].value if col_preco < len(row) else None

            if not nome or not str(nome).strip() or str(nome).strip().upper() in ("NONE", "NAN"):
                continue
            preco = _parse_preco(preco_raw)
            if preco is None:
                continue
            rows_data.append({
                "nome": str(nome).strip(),
                "ean": _parse_ean(ean_raw),
                "preco_base": preco,
            })

        wb_in.close()
        if rows_data:
            logger_local.info(f"_ler_excel_base: {len(rows_data)} produtos (openpyxl fallback)")
        return rows_data

    except Exception as e:
        logger_local.error(f"openpyxl fallback falhou: {e}")
        return []


def _ler_pdf_base(caminho_pdf, progress_callback=None):
    """Lê PDF de tabela do atacadista usando somente extração local."""
    import logging
    logger_local = logging.getLogger(__name__)

    def normalizar_celula(valor):
        return " ".join(str(valor or "").replace("\n", " ").split()).strip()

    diagnostics = []

    def emit_progress(update):
        if not progress_callback:
            return
        try:
            progress_callback(update)
        except Exception as e:
            logger_local.warning(f"callback de progresso falhou: {e}")

    def score_header(headers):
        score = 0
        texto = " ".join(headers)
        if any(k in texto for k in ("PRODUTO", "DESCRI", "ITEM", "MERCADORIA")):
            score += 2
        if "EAN" in texto:
            score += 2
        if any(k in texto for k in ("PRECO", "PREÇO", "VALOR", "R$", "UNIT")):
            score += 2
        return score

    def localizar_colunas(headers):
        col_nome = None
        col_ean = None
        preco_candidates = []

        for i, h in enumerate(headers):
            if col_nome is None and any(k in h for k in ("PRODUTO", "DESCRI", "ITEM", "MERCADORIA")):
                col_nome = i
            if _score_coluna_ean(h) >= 35 or "EAN" in h:
                col_ean = i

            is_price = any(k in h for k in ("PRECO", "PREÇO", "VALOR", "R$", "UNIT"))
            if is_price:
                score = 1
                if any(k in h for k in ("UNIT", "UNITARIO", "UNITÁRIO")):
                    score += 6
                if any(k in h for k in ("TOTAL", "EMB")):
                    score -= 4
                preco_candidates.append((score, i))

        col_preco = max(preco_candidates, default=(0, len(headers) - 1))[1]
        return col_nome if col_nome is not None else 0, col_ean, col_preco

    def linha_ignorada(line):
        upper = line.upper()
        return (
            not line
            or upper.startswith("DADOS DA EMPRESA")
            or upper.startswith("EMPRESA:")
            or upper.startswith("ENDEREÇO:")
            or upper.startswith("ENDERECO:")
            or upper.startswith("CNPJ:")
            or upper.startswith("TELEFONE:")
            or upper.startswith("CLIENTE:")
            or upper.startswith("VENDEDOR:")
            or upper.startswith("REPRES ")
            or upper.startswith("ELEMENTOS QUE")
            or upper.startswith("CÓDIGO ")
            or upper.startswith("CODIGO ")
            or upper.startswith("UN UNIT")
            or upper.startswith("TOTAL ")
            or "DATA EMISS" in upper
            or "DATA CONCLUS" in upper
        )

    def parse_preco_pdf(valor):
        texto = str(valor or "").replace("R$", "").replace(" ", "").strip()
        if not texto:
            return None
        try:
            if "," in texto:
                return float(texto.replace(".", "").replace(",", "."))

            parts = texto.split(".")
            if len(parts) == 2 and len(parts[1]) == 3:
                return float(texto)
            if len(parts) > 2 and len(parts[-1]) == 3:
                return float("".join(parts[:-1]) + "." + parts[-1])
            return float(texto)
        except (TypeError, ValueError):
            return None

    def _ler_texto_bruto(text):
        rows_text = []
        current = None

        row_pattern = _re.compile(
            r"^\s*(?P<code>\d[\w./-]*)\s+(?P<name>.+?)\s+(?P<ean>\d{8,14})\s+(?P<tail>.+)$"
        )
        flexible_row_pattern = _re.compile(
            r"^\s*(?P<name>.+?)\s+(?P<ean>\d{8,14})\s+(?P<tail>.+)$"
        )
        price_pattern = _re.compile(r"(?<!\d)(\d+(?:[.,]\d{2,3})+)(?!\d)")

        for raw_line in str(text or "").splitlines():
            line = normalizar_celula(raw_line)
            if linha_ignorada(line):
                continue

            match = row_pattern.match(line)
            if not match:
                match = flexible_row_pattern.match(line)
            if match:
                prices = price_pattern.findall(match.group("tail"))
                preco = parse_preco_pdf(prices[0]) if prices else None
                if preco is None:
                    current = None
                    continue
                nome = match.group("name").strip()
                nome = _re.sub(r"^\d[\w./-]*\s+", "", nome).strip()
                if len(nome) < 3 or nome.upper() in ("PRODUTO", "CODIGO", "CÓDIGO"):
                    current = None
                    continue
                current = {
                    "nome": nome,
                    "ean": limpar_ean(match.group("ean")),
                    "preco_base": preco,
                }
                rows_text.append(current)
                continue

            if current and not _re.search(r"\d{8,14}", line):
                # Produto quebrado em duas ou mais linhas no PDF.
                if len(line) <= 80 and not price_pattern.search(line):
                    current["nome"] = f"{current['nome']} {line}".strip()

        return rows_text

    def _ler_pdf_por_texto(pdf):
        rows_text = []

        for page_number, page in enumerate(pdf.pages, 1):
            try:
                text = page.extract_text() or ""
            except Exception as e:
                logger_local.warning(f"extract_text falhou na pagina {page_number}: {e}")
                text = ""

            rows_text.extend(_ler_texto_bruto(text))

            if page_number == len(pdf.pages) or page_number % 5 == 0:
                emit_progress({
                    "stage": "extracting_pdf_text",
                    "current_page": page_number,
                    "total_pages": len(pdf.pages),
                    "rows": len(rows_text),
                })

        return rows_text

    rows_data = []
    headers_cache = None

    try:
        from pdfminer.high_level import extract_text

        emit_progress({"stage": "extracting_pdf_text", "rows": 0})
        rows_data = _ler_texto_bruto(extract_text(caminho_pdf) or "")
        diagnostics.append(f"texto_pdfminer_primeiro={len(rows_data)}_produtos")
        emit_progress({"stage": "extracting_pdf_text", "rows": len(rows_data)})
    except Exception as e:
        logger_local.warning(f"leitura inicial por pdfminer falhou: {e}")
        diagnostics.append(f"pdfminer_primeiro_erro={type(e).__name__}")

    try:
        import pdfplumber

        if not rows_data:
            with pdfplumber.open(caminho_pdf) as pdf:
                total_pages = len(pdf.pages)
                diagnostics.append(f"pdfplumber_abriu={total_pages}_paginas")
                emit_progress({"stage": "pdf_opened", "total_pages": total_pages, "rows": 0})

                try:
                    rows_data = _ler_pdf_por_texto(pdf)
                    diagnostics.append(f"texto_pdfplumber_primeiro={len(rows_data)}_produtos")
                except Exception as e:
                    logger_local.warning(f"leitura inicial por texto falhou: {e}")
                    diagnostics.append(f"texto_pdfplumber_primeiro_erro={type(e).__name__}")
                    rows_data = []

                if not rows_data:
                    for page_number, page in enumerate(pdf.pages, 1):
                        try:
                            tables = page.extract_tables()
                        except Exception as e:
                            logger_local.warning(f"extract_tables falhou na pagina {page_number}: {e}")
                            diagnostics.append(f"tables_p{page_number}_erro={type(e).__name__}")
                            tables = []
                        for table in tables:
                            try:
                                if not table or len(table) < 2:
                                    continue

                                header_idx = None
                                best_score = 0
                                for idx, candidate in enumerate(table[:5]):
                                    headers_test = [normalizar_celula(c).upper() for c in candidate]
                                    candidate_score = score_header(headers_test)
                                    if candidate_score > best_score:
                                        best_score = candidate_score
                                        header_idx = idx

                                if header_idx is not None and best_score >= 4:
                                    headers = [normalizar_celula(c).upper() for c in table[header_idx]]
                                    headers_cache = headers
                                    data_rows = table[header_idx + 1:]
                                elif headers_cache:
                                    headers = headers_cache
                                    primeira_linha = " ".join(normalizar_celula(c).upper() for c in table[0])
                                    data_rows = table[1:] if "ELEMENTOS QUE COMPÕEM" in primeira_linha else table
                                else:
                                    continue

                                col_nome, col_ean, col_preco = localizar_colunas(headers)

                                for row in data_rows:
                                    if not row or len(row) <= col_preco:
                                        continue
                                    nome = normalizar_celula(row[col_nome]) if col_nome < len(row) else ""
                                    if not nome or nome.upper() in ("NONE", "", "PRODUTO", "CÓDIGO", "CODIGO"):
                                        continue
                                    ean_raw = normalizar_celula(row[col_ean]) if col_ean is not None and col_ean < len(row) else ""
                                    preco_raw = normalizar_celula(row[col_preco]) if row[col_preco] else ""
                                    preco = parse_preco_pdf(preco_raw)
                                    if preco is None:
                                        continue
                                    rows_data.append({
                                        "nome": nome,
                                        "ean": limpar_ean(ean_raw),
                                        "preco_base": preco,
                                    })
                            except Exception as e:
                                logger_local.warning(f"tabela do PDF ignorada na pagina {page_number}: {e}")
                                continue
                        if page_number == total_pages or page_number % 5 == 0:
                            emit_progress({
                                "stage": "extracting_pdf",
                                "current_page": page_number,
                                "total_pages": total_pages,
                                "rows": len(rows_data),
                            })

                if not rows_data:
                    diagnostics.append("tabelas=0_produtos")
                    emit_progress({"stage": "extracting_pdf_text", "total_pages": total_pages, "rows": 0})
                    try:
                        rows_data = _ler_pdf_por_texto(pdf)
                        diagnostics.append(f"texto_pdfplumber={len(rows_data)}_produtos")
                    except Exception as e:
                        logger_local.warning(f"fallback por texto com pdfplumber falhou: {e}")
                        diagnostics.append(f"texto_pdfplumber_erro={type(e).__name__}")
    except Exception as e:
        logger_local.warning(f"pdfplumber falhou: {e}")
        diagnostics.append(f"pdfplumber_erro={type(e).__name__}")

    if not rows_data:
        try:
            from pdfminer.high_level import extract_text

            emit_progress({"stage": "extracting_pdf_text", "rows": 0})
            rows_data = _ler_texto_bruto(extract_text(caminho_pdf) or "")
            diagnostics.append(f"texto_pdfminer={len(rows_data)}_produtos")
            emit_progress({"stage": "extracting_pdf_text", "rows": len(rows_data)})
        except Exception as e:
            logger_local.warning(f"fallback por pdfminer falhou: {e}")
            diagnostics.append(f"pdfminer_erro={type(e).__name__}")

    if not rows_data:
        detalhe = "; ".join(diagnostics[-8:]) if diagnostics else "sem_diagnostico"
        raise ValueError(
            f"Não encontrei produtos e preços nesse PDF. Tente gerar novamente ou converta o arquivo para Excel (.xlsx). Diagnóstico: {detalhe}."
        )

    return rows_data


def _gerar_excel_de_dados(rows_data, percentuais):
    """Gera Excel com colunas por prazo a partir de dados já extraídos."""
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Tabela de Preços"

    ws_out.cell(1, 1).value = "Tabela de Preços com Prazos — gerada pelo Venpro"
    ws_out.cell(1, 1).font = Font(bold=True, size=11, color="2D2926")

    prazos = [7, 14, 21, 28]
    headers = ["PRODUTO", "EAN"] + [f"{p} dias" for p in prazos]
    header_fill = PatternFill(start_color="B35C44", end_color="B35C44", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    for col_idx, h in enumerate(headers, 1):
        cell = ws_out.cell(3, col_idx)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, item in enumerate(rows_data, 4):
        ws_out.cell(row_idx, 1).value = item["nome"]
        ws_out.cell(row_idx, 2).value = item["ean"]
        for col_offset, prazo in enumerate(prazos, 3):
            pct = percentuais.get(prazo, 0.0)
            preco_final = round(item["preco_base"] * (1 + pct / 100), 2)
            cell = ws_out.cell(row_idx, col_offset)
            cell.value = preco_final
            cell.number_format = '#,##0.00'

    ws_out.column_dimensions['A'].width = 48
    ws_out.column_dimensions['B'].width = 16
    for letra in ['C', 'D', 'E', 'F']:
        ws_out.column_dimensions[letra].width = 11

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb_out.save(tmp.name)
    tmp.close()
    return tmp.name


def gerar_excel_multiprazos(caminho_base, percentuais, progress_callback=None):
    """
    Orquestrador: aceita Excel (.xlsx/.xls) ou PDF e gera tabela com colunas por prazo.
    percentuais: {7: 0.0, 14: 2.5, 21: 4.0, 28: 5.5}
    """
    if caminho_base.lower().endswith('.pdf'):
        rows_data = _ler_pdf_base(caminho_base, progress_callback=progress_callback)
    else:
        rows_data = _ler_excel_base(caminho_base)

    if not rows_data:
        raise ValueError("Nenhum produto encontrado no arquivo enviado")

    if progress_callback:
        try:
            progress_callback({"stage": "writing_excel", "rows": len(rows_data)})
        except Exception:
            pass
    return _gerar_excel_de_dados(rows_data, percentuais)
