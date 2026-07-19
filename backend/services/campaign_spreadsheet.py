from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook


AWARDED_SHEET = "Itens Premiados"
AWARDED_START_COLUMNS = (5, 9, 13, 17, 21, 25)


def _number(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _date_iso(value: Any) -> str | None:
    if isinstance(value, (datetime, date)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    return None


def _rca_code(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value or "").strip()
    return text if text.isdigit() else ""


def parse_campaign_workbook(content: bytes) -> dict:
    """Extrai apenas a apuração necessária; o arquivo original não é persistido."""
    # O arquivo semanal é pequeno; modo normal evita o custo elevado de acessos
    # aleatórios a células que o modo read-only teria neste layout em blocos.
    workbook = load_workbook(BytesIO(content), read_only=False, data_only=True)
    if AWARDED_SHEET not in workbook.sheetnames:
        raise ValueError(f'Aba obrigatória "{AWARDED_SHEET}" não encontrada')

    awarded = workbook[AWARDED_SHEET]
    suppliers = []
    valid_columns = []
    for column in AWARDED_START_COLUMNS:
        name = str(awarded.cell(6, column).value or "").strip()
        goal = _number(awarded.cell(8, column + 1).value)
        if not name or name == "#N/A" or goal <= 0:
            continue
        realized = _number(awarded.cell(7, column + 1).value)
        suppliers.append({
            "name": name,
            "goal": goal,
            "realized": realized,
            "remaining": max(goal - realized, 0),
            "percentage": (realized / goal * 100) if goal else 0,
        })
        valid_columns.append((column, name))

    rca_results: dict[str, dict] = {}
    for row in range(10, awarded.max_row + 1):
        code = _rca_code(awarded.cell(row, 1).value)
        if not code:
            continue
        supplier_results = []
        for column, name in valid_columns:
            supplier_results.append({
                "name": name,
                "prize": _number(awarded.cell(row, column).value),
                "sales": _number(awarded.cell(row, column + 1).value),
                "boxes": _number(awarded.cell(row, column + 2).value),
            })
        rca_results[code] = {
            "code": code,
            "name": str(awarded.cell(row, 2).value or "").strip(),
            "totalPrize": _number(awarded.cell(row, 4).value),
            "awardedItems": supplier_results,
            "industries": {},
        }

    for sheet_name in workbook.sheetnames:
        if sheet_name == AWARDED_SHEET:
            continue
        sheet = workbook[sheet_name]
        header_row = None
        for row in range(1, min(sheet.max_row, 20) + 1):
            if str(sheet.cell(row, 3).value or "").strip().upper() == "RCA":
                header_row = row
                break
        if not header_row:
            continue
        for row in range(header_row + 1, sheet.max_row + 1):
            code = _rca_code(sheet.cell(row, 3).value)
            if not code or code not in rca_results:
                continue
            rca_results[code]["industries"][sheet_name] = {
                "minimumSales": _number(sheet.cell(row, 6).value),
                "sales": _number(sheet.cell(row, 7).value),
                "targetQuantity": _number(sheet.cell(row, 8).value),
                "quantity": _number(sheet.cell(row, 9).value),
            }

    period_end = _date_iso(awarded.cell(4, 10).value)
    if not suppliers:
        raise ValueError("Nenhuma meta válida foi encontrada na aba Itens Premiados")
    if not rca_results:
        raise ValueError("Nenhum RCA válido foi encontrado na aba Itens Premiados")
    return {
        "periodEnd": period_end,
        "suppliers": suppliers,
        "rcaResults": rca_results,
    }
